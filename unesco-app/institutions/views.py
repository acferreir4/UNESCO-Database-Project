import csv
from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic import DetailView, CreateView, UpdateView, DeleteView
from .forms import InstitutionCreateForm
from .models import Institution, ResearchInstituteContact

def exportInstitutionsCsv(request):
    if not request.user.is_staff:
        return redirect('access-denied')

    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="institutions.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Institution', 
        'Continent', 
        'Country', 
        'City',
        'Met',
        'MoC',
        'Status Request',
        'RI 1 Tools',
        'Ethics',
        'Public/Private',
        'Type',
        'General',
        'Role',
        'Student Count',
        'Staff Count',
        'Contact Person(s)',
        'Function',
        'Degree',
        'Contact',
        'Internet Access',
        'Online',
        'Guest Lectures',
        'Environment',
        'Focus on P/S/T',
        'Further',
        'School Size',
        'Community Size',
        'Girl Ratio',
        'Boy Ratio',
        'Qualification',
        'Indigenous',
        'Age',
        'Country',
        'Population',
        'Percent Indigenous',
        'Average Education',
        'Strategy',
        'GDP on Education',
        ])

    for inst in Institution.objects.all():
        contacts = ResearchInstituteContact.objects.filter(institution=inst)
        contact_name = ''
        contact_function = ''
        contact_degree = ''
        contact_email = ''
        if contacts.count() > 0:
            first_contact = contacts.first()
            contact_name = contacts.first().user
            contact_function = contacts.first().function
            contact_degree = contacts.first().degree
            contact_email = contacts.first().user.email

        writer.writerow([
            inst.name, 
            inst.city.country.continent, 
            inst.city.country, 
            inst.city,
            'Yes' if inst.met else 'No',
            'Yes' if inst.moc else 'No',
            'Yes' if inst.status_request else 'No',
            'Yes' if inst.ethics else 'No',
            inst.ri_1_tools,
            'Private' if inst.is_private else 'Public',
            inst.type_of_inst,
            inst.general,
            inst.role,
            inst.student_count,
            inst.staff_count,
            contact_name,
            contact_function,
            contact_degree,
            contact_email,
            inst.internet_access,
            inst.online,
            inst.guest_lectures,
            inst.environment,
            inst.focus_pst,
            inst.further,
            inst.school_size,
            inst.community_size,
            inst.girl_ratio,
            None if inst.girl_ratio is None else 100 - inst.girl_ratio,
            inst.qualifications,
            inst.percent_indigenous,
            inst.age,
            inst.city.country,
            inst.city.country.population,
            inst.city.country.percent_indigenous,
            inst.city.country.average_education,
            inst.city.country.strategy,
            inst.city.country.percent_gdp_on_ed,
            ])

        if contacts.count() > 1:
            for contact in ResearchInstituteContact.objects.filter(institution=inst).exclude(user=first_contact.user):
                writer.writerow(['', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                    contact.user,
                    contact.function,
                    contact.degree,
                    contact.user.email,
                    '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])

    return response

def exportInstitutionsXlsx(request):
    if not request.user.is_staff:
        return redirect('access-denied')

    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="institutions.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Institutions'

    columns = [
        'Institution', 
        'Continent', 
        'Country', 
        'City',
        'Met',
        'MoC',
        'Status Request',
        'RI 1 Tools',
        'Ethics',
        'Public/Private',
        'Type',
        'General',
        'Role',
        'Student Count',
        'Staff Count',
        'Contact Person(s)',
        'Function',
        'Degree',
        'Contact',
        'Internet Access',
        'Online',
        'Guest Lectures',
        'Environment',
        'Focus on P/S/T',
        'Further',
        'School Size',
        'Community Size',
        'Girl Ratio',
        'Boy Ratio',
        'Qualification',
        'Indigenous',
        'Age',
        'Country',
        'Population',
        'Percent Indigenous',
        'Average Education',
        'Strategy',
        'GDP on Education',
        ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for inst in Institution.objects.all():
        row_num += 1

        contacts = ResearchInstituteContact.objects.filter(institution=inst)
        contact_name = ''
        contact_function = ''
        contact_degree = ''
        contact_email = ''
        if contacts.count() > 0:
            first_contact = contacts.first()
            contact_name = contacts.first().user.username
            contact_function = contacts.first().function
            contact_degree = contacts.first().degree
            contact_email = contacts.first().user.email

        row = [
            inst.name, 
            inst.city.country.continent, 
            inst.city.country.name, 
            inst.city.name,
            'Yes' if inst.met else 'No',
            'Yes' if inst.moc else 'No',
            'Yes' if inst.status_request else 'No',
            'Yes' if inst.ethics else 'No',
            inst.ri_1_tools,
            'Private' if inst.is_private else 'Public',
            inst.type_of_inst,
            inst.general,
            inst.role,
            inst.student_count,
            inst.staff_count,
            contact_name,
            contact_function,
            contact_degree,
            contact_email,
            inst.internet_access,
            inst.online,
            inst.guest_lectures,
            inst.environment,
            inst.focus_pst,
            inst.further,
            inst.school_size,
            inst.community_size,
            inst.girl_ratio,
            None if inst.girl_ratio is None else 100 - inst.girl_ratio,
            inst.qualifications,
            inst.percent_indigenous,
            inst.age,
            inst.city.country.name,
            inst.city.country.population,
            inst.city.country.percent_indigenous,
            inst.city.country.average_education,
            inst.city.country.strategy,
            inst.city.country.percent_gdp_on_ed,
            ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

        if contacts.count() > 1:
            for contact in ResearchInstituteContact.objects.filter(institution=inst).exclude(user=first_contact.user):
                row_num += 1
                row = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
                    contact.user.username,
                    contact.function,
                    contact.degree,
                    contact.user.email,
                    '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value

    workbook.save(response)
    return response


def manageInstitutions(request):
    context = {
        'institutions': Institution.objects.all()
    }
    return render(request, 'institutions/manage_institution.html', context)

class InstitutionDetailView(DetailView):  # looks for <institutions/institution_detail.html
    model = Institution

class InstitutionCreateView(LoginRequiredMixin, SuccessMessageMixin, UserPassesTestMixin, CreateView): 
    template_name = 'institutions/institution_form.html'
    form_class = InstitutionCreateForm
    success_url = '/manage-institutions/'
    success_message = "%(name)s has been created!"

    def form_valid(self, form):
        return super().form_valid(form)

    def test_func(self):
        return self.request.user.is_staff


class InstitutionUpdateView(LoginRequiredMixin, SuccessMessageMixin, UserPassesTestMixin, UpdateView):
    template_name = 'institutions/institution_form.html'
    form_class = InstitutionCreateForm
    success_message = "%(name)s has been updated!"

    def form_valid(self, form):
        return super().form_valid(form)

    def test_func(self):
        return self.request.user.is_staff

    def get_queryset(self):
        return Institution.objects.all()


class InstitutionDeleteView(LoginRequiredMixin, SuccessMessageMixin, UserPassesTestMixin, DeleteView):
    model = Institution
    success_url = '/manage-institutions/'
    success_message = "Institution deleted."

    def test_func(self):
        return self.request.user.is_staff



# Create your views here.
#def createInstitution(request):
#    if request.method == 'POST':
#        form = InstitutionCreateForm(request.POST)
#        if form.is_valid():
#            form.save()
#            name = form.cleaned_data.get('name')
#            messages.success(request, f'Institution %s has been created!' % name)
#            return redirect('manageInstitutions')
#    else:
#        form = InstitutionCreateForm()
#    return render(request, 'institutions/create_institution.html', {'form':form})

#def institutionProfile(request):
#    return render(request, 'institutions/institution_profile.html')

