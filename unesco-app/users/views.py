import csv
import random
import string
from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.template.loader import render_to_string
from .forms import UserRegisterForm, UserUpdateForm, ProfileUpdateForm
from .models import User

@login_required 
def register(request):
    if not request.user.is_staff:
        return redirect('access-denied')
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            temp_pass = randomPassword()
            form.cleaned_data['password1'] = temp_pass
            form.cleaned_data['password2'] = temp_pass
            request.user._temp_pass = temp_pass
            form.save()
            
            username = form.cleaned_data.get('username')
            email = form.cleaned_data.get('email')
            name = f'%s %s' % (form.cleaned_data.get('first_name'), form.cleaned_data.get('last_name'))
            role = form.cleaned_data.get('role')
            send_registration_email(username, email, name, role, temp_pass)

            messages.success(request, f'Account for user %s has been created! An email has been sent to %s with instructions to change their password.' % (username, email))
            return redirect('register')
    else:
        form = UserRegisterForm()
    return render(request, 'users/register.html', {'form':form})

def randomPassword(stringLength=10):
    """Generate a random string of fixed length."""
    code = string.ascii_letters + string.digits
    return ''.join(random.choice(code) for i in range(stringLength))

def send_registration_email(username, email, name, role, temp_password):
    subject = f'Welcome %s to UNESCO Indigenous Research!' % name
    send_mail(
            subject = subject,
            message = '',
            from_email = '',
            recipient_list = [email],
            html_message = render_to_string(
                'users/registration_email.html', {'username': username, 'role': role, 'temp_password': temp_password}
                ),
            fail_silently=False,
            )

@login_required
def profile(request):
    if request.method == 'POST':
        u_form = UserUpdateForm(request.POST, instance=request.user)
        p_form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user.profile)
        if u_form.is_valid() and p_form.is_valid():
            u_form.save()
            p_form.save()
            messages.success(request, f'Your contact information has been updated!')
            return redirect('profile')
    else:
        u_form = UserUpdateForm(instance=request.user)
        p_form = ProfileUpdateForm(instance=request.user.profile)

    context = {
            'u_form': u_form,
            'p_form': p_form
    }

    return render(request, 'users/profile.html', context)

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password was successfully updated!')
            return redirect('change-password')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'users/change_password.html', {'form':form})

def accessDenied(request):
    return render(request, 'users/access_denied.html')


def exportUsersCsv(request):
    if not request.user.is_staff:
        return redirect('access-denied')
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="users.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Username', 
        'First Name', 
        'Last Name', 
        'Email',
        'Institution',
        'Role',
        'Admin',
        'Active',
        'Date Registered',
        'Last Login',
        ])

    for user in User.objects.order_by('institution'):
        writer.writerow([
            user.username, 
            user.first_name, 
            user.last_name, 
            user.email,
            user.institution,
            user.role,
            'Yes' if user.is_staff else 'No',
            'Yes' if user.is_active else 'No',
            user.date_joined,
            user.last_login,
            ])

    return response

def exportUsersXlsx(request):
    if not request.user.is_staff:
        return redirect('access-denied')
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="users.xlsx"'

    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Users'

    #Define the titles for columns
    columns = [
        'Username',
        'First name',
        'Last name',
        'Email',
        'Institution',
        'Role',
        'Admin',
        'Active',
        'Date Registered',
        'Last Login',
        ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for user in User.objects.order_by('institution'):
        row_num += 1

        # Define the data for each cell in the row
        row = [
            user.username, 
            user.first_name, 
            user.last_name, 
            user.email,
            user.institution.name,
            user.role,
            'Yes' if user.is_staff else 'No',
            'Yes' if user.is_active else 'No',
            user.date_joined,
            user.last_login,
            ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
